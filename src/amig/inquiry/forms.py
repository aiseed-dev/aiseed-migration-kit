"""申込様式 xlsx の生成(様式プロファイルからの派生物の一つ)。

様式の唯一の定義は様式プロファイル(AsciiDoc。DESIGN.md §5)で、xlsx は
そこから生成される記入チャネル。名前付きセルの座標はこのモジュールに
閉じ、他の場所でハードコードしない(読み取り parse.py もここの names() を
参照する)。Excel の定義名にハイフンは使えないため、定義名のみ
アンダースコアを使う。

様式は公開サイトからダウンロードできる(build が dist/forms/ に載せる)。
受付アドレスはページ本文に書かず、様式の中にだけ置く(ボット収集対策。
DESIGN.md §5)。発行キーは埋め込まない。
"""

import hashlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from amig.site import FormDef, Site

# 様式の版(送信テキストのプロトコル版)。定義名の体系が変わったら上げる。
# 2 = 様式プロファイル移行(定義名が f_<key> から f_c<N> に変わった)
FORM_VER = 2
SHEET = "記入シート"
TEXT_SHEET = "送信用テキスト"

_STAFF_ROW = 4
_FIELD_ROW0 = 6  # 記入欄の先頭行(1欄1行)
STAFF_LABEL = "宛先(担当)"

# 送信用テキストの固定項目名(→ parse.parse_text が同じ表で読む)
KEY_KIND = "様式"
KEY_VER = "様式版"
KEY_SITE = "受付"  # site.yaml 由来の識別子(偽様式対策。受付側で照合。§11)
KEY_REV = "様式指紋"  # 項目構成のハッシュ(下記 rev()。版の上げ忘れ対策)
KEY_STAFF = "宛先"


def rev(form: FormDef) -> str:
    """様式指紋: 項目構成(ラベル+制約の並び)から決定的に導出する短い
    ハッシュ。プロファイルの項目を挿入・入替すると必ず変わるため、
    FORM_VER(人が上げるプロトコル版)と違い、配布済み xlsx の位置ずれ
    (f_c1.. が別の項目を指す)を機械的に検出できる。"""
    src = form.key + "\n" + "\n".join(f"{f.label}|{f.spec.sql}" for f in form.fields)
    return hashlib.sha256(src.encode("utf-8")).hexdigest()[:8]


def names(form: FormDef) -> dict[str, str]:
    """定義名 → セル座標(様式の機械可読部の全リスト)。"""
    n = {
        "form_kind": "H1",
        "form_ver": "H2",
        "form_site": "H3",
        "form_rev": "H4",
        "staff": f"C{_STAFF_ROW}",
    }
    for i, f in enumerate(form.fields):
        n[f"f_{f.key}"] = f"C{_FIELD_ROW0 + i}"
    return n


def text_keys(form: FormDef) -> dict[str, str]:
    """定義名 → 送信用テキストの項目名(メール本文の `項目: 値` の左側)。"""
    keys = {
        "form_kind": KEY_KIND,
        "form_ver": KEY_VER,
        "form_site": KEY_SITE,
        "form_rev": KEY_REV,
        "staff": KEY_STAFF,
    }
    for f in form.fields:
        keys[f"f_{f.key}"] = f.label
    return keys


_FILL_INPUT = PatternFill("solid", fgColor="FFFDE7")  # 記入欄(薄い黄色)
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _draw(ws: Worksheet, site: Site, form: FormDef, note: str) -> int:
    """レイアウトを描き、案内文の行番号を返す。

    E 列は項目の補足(様式プロファイルの自由記述)。機械は読まない。
    """
    ws["A1"] = form.label
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = site.title
    ws[f"A{_STAFF_ROW}"] = STAFF_LABEL
    staff_cell = ws[f"C{_STAFF_ROW}"]
    staff_cell.fill = _FILL_INPUT
    staff_cell.border = _BORDER
    for i, f in enumerate(form.fields):
        row = _FIELD_ROW0 + i
        ws[f"A{row}"] = f.label + ("" if f.required else "(任意)")
        cell = ws[f"C{row}"]
        cell.fill = _FILL_INPUT
        cell.border = _BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        hints = []
        if f.spec.choices:
            hints.append(" / ".join(f.spec.choices) + " から選択")
        if f.note:
            hints.append(f.note.replace("\n", " "))
        if hints:
            e = ws[f"E{row}"]
            e.value = " ".join(hints)
            e.font = Font(size=9, color="666666")
            e.alignment = Alignment(wrap_text=True, vertical="top")
    note_row = _FIELD_ROW0 + len(form.fields) + 1
    ws[f"A{note_row}"] = note
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["E"].width = 30
    ws.print_area = f"A1:F{note_row}"
    ws.page_setup.fitToWidth = 1
    return note_row


def build(site: Site, form: FormDef, submit_addr: str) -> Workbook:
    """様式を組み立てる(1枚目=記入シート、2枚目=送信用テキスト)。

    submit_addr は受付アドレス。様式にだけ書き、公開ページの本文には載せない。
    """
    note = (
        f"記入後、「{TEXT_SHEET}」シートの本文をコピーして {submit_addr} へ"
        "メール本文として送信してください。このファイルを添付して"
        "送っていただいても構いません。印刷して FAX でも受け付けます。"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    _draw(ws, site, form, note)

    # 機械可読メタ(印刷範囲外の列に置き、非表示にする)
    n = names(form)
    ws[n["form_kind"]] = form.key
    ws[n["form_ver"]] = FORM_VER
    ws[n["form_site"]] = site.name  # 識別子(偽様式対策。受付側で照合)
    ws[n["form_rev"]] = rev(form)  # 様式指紋(項目構成の変更検出)
    ws.column_dimensions["H"].hidden = True

    for name, coord in n.items():
        ref = f"{quote_sheetname(SHEET)}!${coord[0]}${coord[1:]}"
        wb.defined_names[name] = DefinedName(name, attr_text=ref)

    # 宛先(担当)はドロップダウン
    _dropdown(ws, n["staff"], [s.label for s in site.staff])

    # 選択肢(check in)のある項目もドロップダウン(制約から派生)
    for f in form.fields:
        if f.spec.choices:
            _dropdown(ws, n[f"f_{f.key}"], list(f.spec.choices))

    # 入力セル以外はシート保護
    meta = ("form_kind", "form_ver", "form_site", "form_rev")
    for name, coord in n.items():
        if name not in meta:
            ws[coord].protection = Protection(locked=False)
    ws.protection.sheet = True

    _text_sheet(wb, form, submit_addr)
    return wb


def _dropdown(ws: Worksheet, coord: str, values: list[str]) -> None:
    """ドロップダウン(インラインリスト)を張る。

    Excel のインラインリストはカンマ区切り・二重引用符不可・255文字上限。
    表せない選択肢(カンマや " を含む・長すぎる)はドロップダウンを付けず
    自由入力にする——選択肢の案内は E 列に出ており、正の検証は常に
    サーバー側(選択肢を分裂・欠落させて検証と食い違うより良い)。
    """
    joined = ",".join(values)
    if any("," in v or '"' in v for v in values) or len(joined) > 255:
        return
    dv = DataValidation(
        type="list",
        formula1=f'"{joined}"',
        allow_blank=True,
        showErrorMessage=True,
        error="リストから選んでください",
    )
    ws.add_data_validation(dv)
    dv.add(ws[coord])


TEXT_ROW0 = 6  # 送信用テキストの本文が始まる行(A列。数式で自動生成)


def _text_sheet(wb: Workbook, form: FormDef, submit_addr: str) -> None:
    """「送信用テキスト」シート。

    A列: 数式で本文を1行1項目で自動生成(Excel・LibreOffice・OnlyOffice の
    どれでも動く)。行を範囲選択してコピー→メール本文に貼るだけ。
    C列: 内蔵マクロ(OnlyOffice 専用)が一括コピー用に1セルへまとめる補助。
    """
    ws = wb.create_sheet(TEXT_SHEET)
    ws["A1"] = "送信用テキスト(メール本文に貼り付けて送る)"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = (
        "おすすめ: 無料の OnlyOffice で開き、マクロ「送信用テキスト」を"
        "実行すると、未記入のチェック付きで C 列に本文がまとまります。"
        f"コピーして {submit_addr} へメール本文として送信してください。"
    )
    ws["A3"] = (
        "Excel 等では下の A 列に本文が自動で出ます(チェックなし)。"
        f"{TEXT_ROW0}行目から下を選択してコピーしてください。"
        "このファイルをそのまま添付して送っていただいても構いません。"
    )
    ws[f"A{TEXT_ROW0 - 1}"] = "↓ ここから下をコピー"
    ws[f"A{TEXT_ROW0 - 1}"].font = Font(bold=True)
    keys = text_keys(form)
    for i, (name, coord) in enumerate(names(form).items()):
        ref = f"'{SHEET}'!{coord}"
        ws.cell(
            row=TEXT_ROW0 + i,
            column=1,
            value=f'=IF({ref}="","","{keys[name]}: "&{ref})',
        )
    ws.column_dimensions["A"].width = 90


def macro_js(site: Site, form: FormDef) -> str:
    """様式内蔵マクロ(OnlyOffice JavaScript)を様式定義から生成する。

    本文の生成自体は A 列の数式が担う(全表計算ソフト共通)。マクロは
    OnlyOffice 利用者向けの補助で、未記入チェックのうえ C 列の1セルに
    本文をまとめて一括コピーしやすくする。
    出力: `amig macro <site> <form>`。チェック定義はサーバー側の検証
    (parse)と同じ様式プロファイルから生成されるため、様式を変えたら再生成する。
    """
    n = names(form)
    keys = text_keys(form)
    pairs = ",\n    ".join(f'["{keys[name]}", "{coord}"]' for name, coord in n.items())
    required = ",\n    ".join(
        f'["{f.label}", "{n[f"f_{f.key}"]}"]' for f in form.fields if f.required
    )
    return f"""// 様式マクロ({form.label}): 未記入チェックのうえ、送信用テキストを
// 「{TEXT_SHEET}」シートに書き出す。OnlyOffice(Desktop / Docs)専用。
// このファイルは自動生成(amig macro)——手で直さず、様式プロファイル(.adoc)を直して再生成する
// チェック定義はサーバー側の検証(parse)と同じ様式プロファイルから生成される
(function () {{
  var src = Api.GetSheet("{SHEET}");
  var out = Api.GetSheet("{TEXT_SHEET}");
  function val(coord) {{
    var v = src.GetRange(coord).GetValue();
    return v === null ? "" : String(v).trim();
  }}

  // 1) 未記入チェック(正の検証はサーバー側。ここは送信前の親切)
  var issues = [];
  if (val("{n["staff"]}") === "") issues.push("「{STAFF_LABEL}」が未選択です");
  var required = [
    {required}
  ];
  required.forEach(function (kv) {{
    if (val(kv[1]) === "") issues.push("「" + kv[0] + "」が未記入です");
  }});
  if (issues.length) {{
    out.GetRange("C{TEXT_ROW0 - 1}")
      .SetValue("【未記入があります。本文は作成されませんでした】");
    out.GetRange("C{TEXT_ROW0}").SetValue(issues.join("\\n"));
    out.SetActive();
    return;
  }}

  // 2) 送信用テキストの生成(1セルに一括コピー用)
  var map = [
    {pairs}
  ];
  var lines = [];
  map.forEach(function (kv) {{
    var v = val(kv[1]);
    if (v !== "") lines.push(kv[0] + ": " + v);
  }});
  out.GetRange("C{TEXT_ROW0 - 1}").SetValue("一括コピー用(マクロ出力)");
  out.GetRange("C{TEXT_ROW0}").SetValue(lines.join("\\n"));
  out.SetActive();
}})();
"""
