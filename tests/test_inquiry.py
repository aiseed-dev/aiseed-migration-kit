"""inquiry: 様式 xlsx の生成 → 読み取りの往復と、送信用テキストの寛容パーサ。

様式の定義は sites/example/forms/*.adoc(様式プロファイル)。
"""

import io

import pytest
from openpyxl import load_workbook

from amig.inquiry import forms, parse

FILLED = {
    "ご所属(企業・団体名)": "株式会社○○製作所",
    "お名前": "山田 太郎",
    "電話番号": "088-000-0000",
    "メールアドレス": "taro@example.co.jp",
    "ご用件": "引張試験について相談したい",
}


def _xlsx(example, staff_label="総合窓口", ver=None, values=FILLED) -> bytes:
    """example の contact 様式を組み立てて記入し、bytes で返す。"""
    form = example.form("contact")
    wb = forms.build(example, form, "uketsuke@example.jp")
    ws = wb[forms.SHEET]
    n = forms.names(form)
    if staff_label:
        ws[n["staff"]] = staff_label
    for label, v in values.items():
        f = form.field(label)
        ws[n[f"f_{f.key}"]] = v
    if ver is not None:
        ws[n["form_ver"]] = ver
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_roundtrip(example):
    inq = parse.parse_xlsx(_xlsx(example), example)
    assert inq.form.key == "contact"
    assert inq.staff.key == "general"
    assert inq.values["お名前"] == "山田 太郎"


def test_xlsx_missing_required(example):
    values = dict(FILLED)
    del values["電話番号"]
    with pytest.raises(parse.Invalid) as e:
        parse.parse_xlsx(_xlsx(example, values=values), example)
    assert any("電話番号" in s for s in e.value.issues)


def test_xlsx_no_staff(example):
    with pytest.raises(parse.Invalid) as e:
        parse.parse_xlsx(_xlsx(example, staff_label=""), example)
    assert any("宛先" in s for s in e.value.issues)


def test_xlsx_wrong_version(example):
    with pytest.raises(parse.Invalid, match="版"):
        parse.parse_xlsx(_xlsx(example, ver=99), example)


def test_not_a_form(example):
    with pytest.raises(parse.Invalid, match="様式のファイル"):
        parse.parse_xlsx(b"not an xlsx", example)


def test_workbook_structure(example):
    """様式の作り: シート保護・送信用テキスト・非表示メタ。"""
    wb = load_workbook(io.BytesIO(_xlsx(example)))
    ws = wb[forms.SHEET]
    assert ws.protection.sheet
    assert wb[forms.TEXT_SHEET]["A1"].value.startswith("送信用テキスト")
    assert ws.column_dimensions["H"].hidden


def test_choices_dropdown(example):
    """選択肢(check in)は制約からドロップダウンに派生する。"""
    form = example.form("shiken")
    wb = forms.build(example, form, "uketsuke@example.jp")
    ws = wb[forms.SHEET]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("引張試験" in f for f in formulas)


def _text(example, values=FILLED, kind="contact") -> str:
    lines = [
        f"{forms.KEY_KIND}: {kind}",
        f"{forms.KEY_VER}: {forms.FORM_VER}",
        "宛先: 総合窓口",
    ]
    for label, v in values.items():
        lines.append(f"{label}: {v}")
    return "\n".join(lines)


def test_text_roundtrip(example):
    inq = parse.parse_text(_text(example), example)
    assert inq is not None
    assert inq.staff.key == "general"
    assert inq.values["ご所属(企業・団体名)"] == "株式会社○○製作所"


def test_text_is_lenient(example):
    body = _text(example)
    # 返信の引用・全角コロン・前後の空白に耐える
    quoted = "\n".join("> " + line.replace(": ", ": ", 1) for line in body.splitlines())
    inq = parse.parse_text(quoted, example)
    assert inq is not None
    assert inq.values["お名前"] == "山田 太郎"


def test_text_not_a_form(example):
    assert parse.parse_text("こんにちは。見積をお願いします。", example) is None


def test_text_unknown_kind(example):
    with pytest.raises(parse.Invalid):
        parse.parse_text(
            f"様式: nai\n様式版: {forms.FORM_VER}\n", example
        )


def test_staff_prefix_match(example):
    body = _text(example).replace("宛先: 総合窓口", "宛先: 総合窓口(平日)")
    inq = parse.parse_text(body, example)
    assert inq is not None and inq.staff.key == "general"


SHIKEN = {
    "企業・団体名": "株式会社○○製作所",
    "ご担当者名": "山田 太郎",
    "電話番号": "088-000-0000",
    "メールアドレス": "taro@example.co.jp",
    "試験項目": "引張試験",
    "数量": "3",
}


def test_constraint_email_pattern(example):
    values = dict(FILLED, メールアドレス="アットマークなし")
    with pytest.raises(parse.Invalid) as e:
        parse.parse_text(_text(example, values), example)
    assert any("メールアドレス" in s for s in e.value.issues)


def test_repair_zenkaku_integer(example):
    """全角数字は NFKC 修復で半角になり、integer 制約を通る。"""
    values = dict(SHIKEN, 数量="3")
    inq = parse.parse_text(_text(example, values, kind="shiken"), example)
    assert inq is not None
    assert inq.values["数量"] == "3"


def test_repair_wareki_date(example):
    """和暦は date 制約への決定的な修復で ISO になる。"""
    values = dict(SHIKEN, 希望日="令和8年8月1日")
    inq = parse.parse_text(_text(example, values, kind="shiken"), example)
    assert inq is not None
    assert inq.values["希望日"] == "2026-08-01"


def test_constraint_between(example):
    values = dict(SHIKEN, 数量="101")
    with pytest.raises(parse.Invalid) as e:
        parse.parse_text(_text(example, values, kind="shiken"), example)
    assert any("数量" in s for s in e.value.issues)


def test_constraint_choices(example):
    values = dict(SHIKEN, 試験項目="爆発試験")
    with pytest.raises(parse.Invalid) as e:
        parse.parse_text(_text(example, values, kind="shiken"), example)
    assert any("試験項目" in s for s in e.value.issues)


def test_macro_js(example):
    js = forms.macro_js(example, example.form("contact"))
    assert "電話番号" in js  # 必須チェックが様式定義から生成される
    assert forms.TEXT_SHEET in js
    assert "Api.GetSheet" in js


def test_dropdown_skipped_for_comma_choice():
    """カンマ入り選択肢はインラインリストで表せないためドロップダウンを
    付けない(選択肢が分裂してサーバー検証と食い違うより良い)。"""
    from types import SimpleNamespace

    from amig import site as site_mod
    from amig.inquiry import spec

    dummy = SimpleNamespace(
        name="x", title="t", staff=(site_mod.Staff(key="g", label="窓口", folder="staff/g"),)
    )
    prof = site_mod.FormDef(
        key="a",
        label="様式",
        intro="",
        fields=(
            site_mod.Field(
                key="c1",
                label="区分",
                spec=spec.parse(
                    "区分", "varchar(20), check (区分 in ('1,000本まで', '継続'))"
                ),
            ),
        ),
    )
    wb = forms.build(dummy, prof, "a@example.jp")
    ws = wb[forms.SHEET]
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert not any("000本まで" in f for f in formulas)
    assert any("窓口" in f for f in formulas)  # 宛先側は通常どおり


def test_fingerprint_changes_on_reorder(example):
    """項目の挿入・入替で様式指紋が必ず変わる(位置ずれの検出根拠)。"""
    form = example.form("contact")
    r1 = forms.rev(form)
    from amig.site import FormDef

    swapped = FormDef(
        key=form.key,
        label=form.label,
        intro=form.intro,
        fields=(form.fields[1], form.fields[0], *form.fields[2:]),
    )
    assert forms.rev(swapped) != r1


def test_stale_fingerprint_rejected(example):
    """指紋が違う(構成が変わった)様式からの投稿は最新様式へ案内。"""
    body = _text(example) + f"\n{forms.KEY_REV}: 00000000"
    with pytest.raises(parse.Invalid, match="更新"):
        parse.parse_text(body, example)


def test_current_fingerprint_accepted(example):
    form = example.form("contact")
    body = _text(example) + f"\n{forms.KEY_REV}: {forms.rev(form)}"
    inq = parse.parse_text(body, example)
    assert inq is not None
